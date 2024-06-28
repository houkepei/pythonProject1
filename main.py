from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def read_excel_hyperlinks(file_path):
    """
    读取Excel文件中指定工作表的超链接信息。

    参数:
    file_path: string, Excel文件的路径。

    返回:
    dict, 包含超链接信息的字典。字典的键是工作表名称，值是另一个字典，
    其中键是单元格坐标，值是包含超链接URL和显示文本的字典。
    """
    """读取Excel文件中的超链接"""
    # 以只读模式加载工作簿，只获取数据，不解析公式
    workbook = load_workbook(filename=file_path, data_only=True)  # data_only=True确保获取公式计算的结果
    # 初始化一部字典来存储所有超链接信息
    hyperlinks_dict = {}
    # 定义目标工作表名称
    sheet_name = '目录'
    # 检查目标工作表是否存在
    if sheet_name in workbook.sheetnames:
        # 获取目标工作表
        sheet = workbook[sheet_name]
        # 初始化一部字典来存储当前工作表的超链接信息
        hyperlinks_in_sheet = {}
        # 遍历工作表中的每一行
        for row in sheet.iter_rows():
            # 遍历行中的每个单元格
            for index, cell in enumerate(row):
                # 检查单元格是否有超链接
                if cell.hyperlink:
                    # 提取超链接的URL和显示文本
                    hyperlink_info = {
                        # "url": cell.hyperlink.target,
                        "url": cell.value,
                        "display_text": row[index + 1].value,
                    }
                    # 如果当前工作表还未在字典中记录，则创建新条目
                    if sheet_name not in hyperlinks_in_sheet:
                        hyperlinks_in_sheet[cell.coordinate] = hyperlink_info
                    # 如果当前工作表已存在条目，则将新超链接信息追加到列表中
                    else:
                        hyperlinks_in_sheet[cell.coordinate].append(hyperlink_info)
        # 将当前工作表的超链接信息添加到总字典中
        hyperlinks_dict[sheet_name] = hyperlinks_in_sheet
    # 返回包含所有超链接信息的字典
    return hyperlinks_dict


def process_linked_files(input_excel_path, output_excel_path):
    """
    处理包含超链接的Excel文件，并合并链接的Excel表格数据。

    参数:
    input_excel_path: str - 输入Excel文件的路径，包含超链接。
    output_excel_path: str - 输出Excel文件的路径，用于保存合并后的数据。

    返回:
    无返回值，但会生成一个新的Excel文件，其中包含从所有链接的Excel文件中合并的数据。
    """
    # 创建一个新的工作簿，用于存储合并后的数据
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    """处理Excel文件中的超链接，并进一步处理链接的Excel文件"""

    # 读取输入Excel文件中的超链接，并将其存储在一部字典中
    """读取Excel文件中的超链接"""
    hyperlinks_dict = read_excel_hyperlinks(input_excel_path)

    # 遍历字典中的每个工作表和超链接
    for sheet_name, links in hyperlinks_dict.items():
        print(f"从 sheet: {sheet_name}，获取的表名")
        for coord, link_info in links.items():
            link_url = link_info["url"]
            display_text = link_info["display_text"]
            print(f"打开表  {coord} pointing to {link_url}")

            try:
                # 加载链接的Excel文件，只读取数据，不处理公式
                workbook = load_workbook(filename=input_excel_path, data_only=True)
                # 获取链接指向的工作表
                # 对linked_df进行任何必要的处理，比如筛选、转换等
                sheet = workbook[link_url]
                # 处理 sheet，并合并到新的文件
                # 在此处添加对sheet的具体处理逻辑
                # 以下仅为示例，实际处理逻辑根据需求定制
                # 遍历链接工作表中的数据行，从第三行开始，跳过标题行
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    # 如果遇到空行，停止当前循环
                    # 假设需要提取B到G列的值
                    if row[1] is None:
                        break
                    # 提取B到G列的数据
                    columns_B_to_G = row[1:7]  # 这将提取第2到第7个元素，即B到G列的值
                    # 在数据前插入显示文本和链接地址
                    # 如果你想在指定位置插入
                    columns_B_to_G_list = list(columns_B_to_G)
                    columns_B_to_G_list.insert(0, display_text)
                    columns_B_to_G_list.insert(0, link_url)
                    # 将修改后的数据转换回元组
                    # 将列表再转换回元组
                    columns_B_to_G = tuple(columns_B_to_G_list)

                    # 打印或进一步处理提取的数据
                    # 打印或进一步处理这些值
                    print(columns_B_to_G)
                    # 将数据添加到合并的工作表中
                    # 假设处理后的数据直接追加到merged_ws
                    merged_ws.append(columns_B_to_G)
            except Exception as e:
                # 如果处理过程中发生异常，打印错误信息
                print(f"处理文件异常 {link_url}: {e}")

    # 保存合并后的数据到新的Excel文件
    # 所有工作表处理完毕后，保存合并的数据到新文件
    merged_wb.save(output_excel_path)


if __name__ == '__main__':
    # 示例用法
    # a = input("请输入文件路径和名称：")
    a = 'example_input.xlsx'
    print("你输入的内容是：" + a)
    b = 'processed_example.xlsx'
    process_linked_files(a, b)
